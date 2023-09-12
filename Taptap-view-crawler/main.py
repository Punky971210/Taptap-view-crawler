import random
import time
import datetime

import openpyxl as xl
import collections

from selenium import webdriver
from selenium.webdriver import ActionChains

Game_Dict = []
Game_Dict.append({"id": 61620, "name": "决战平安京"})
Game_Dict.append({"id": 2301, "name": "王者荣耀"})

# "https://www.taptap.cn/app/61620/review?order=default&page=&sort=new"
url_head = "https://www.taptap.cn/app/"  # url开头
url_tail = "/review?order=default&page=&sort=new"

game_id = str(Game_Dict[0]["id"])
game_name = Game_Dict[0]["name"]

file_name_head = "_comment_count.xlsx"
file_name = game_name + file_name_head
url_full = url_head + game_id + url_tail

page_total = 200

comment_out = []  # 输出容器
score_out = []
time_out = []
user_out = []
date_out = []
collapse_out = []

label_R = ["user", "comment", "score", "time", "collapse", "date"]
label_C = ["date", "star-1", "star-2", "star-3", "star-4", "star-5",
           "备注：统计包含被折叠评论,截取时长不足一天的不统计"]

date_present = []

option = webdriver.ChromeOptions()
option.add_argument('disable-infobars')
# 配置ChromeOptions
# option.add_argument("--proxy-server=http://127.0.0.1:8888")
option.add_experimental_option('excludeSwitches', ['enable-automation'])
# option.add_argument('headless')  # 添加了以后就不会弹出浏览器

browser = webdriver.Chrome(options=option)
browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": """
    Object.defineProperty(navigator, 'webdriver', {
      get: () => undefined
    })
  """
})


def get_block(url):
    browser.get(url)
    time.sleep(5)
    browser.refresh()

    try:
        collect_start = browser.find_element(
            by="class name", value="review-item--in-app-tab__content.review-item__content"
        )
        if collect_start:
            print("开始收集")
    except:
        print("未找到评论或未进入正确页面，请检查！")
        pass

    for a in range(page_total):
        time.sleep(random.randint(2, 4))
        js = "window.scrollTo(0,document.body.scrollHeight)"
        browser.execute_script(js)
        try:
            # browser.find_element(by="partial link text",
            #                      value="该评价被管理员折叠").click()

            collapse = browser.find_elements(by="class name",
                                             value="review-item__collapsed.paragraph-m14-w14.gray-04.flex-center")
            for fold in collapse:
                ActionChains(browser).move_to_element(fold).click().perform()
                print("展开一个被折叠页面")
        except:
            pass
        print("正在收集 第" + str(a + 1) + "页")

    time.sleep(random.randint(2, 4))
    search_blocks = browser.find_elements(
        by="class name", value="review-item--in-app-tab__content.review-item__content"
    )  # 变量名中有“空格”用“.”替换

    # print(browser.page_source)
    # print(search_blocks)
    for block in search_blocks:
        print(block)

        user_out.append(get_user(block))
        score_out.append(get_score(block))
        comment_out.append(get_comment(block))

        time_text = get_time_text(block)
        time_out.append(time_text)
        post_time = get_time(time_text)

        try:
            date_text = post_time.strftime('%m-%d')
            date_out.append(date_text)
        except:
            date_out.append(post_time)

        collapse_out.append(is_collapsed(block))

    return search_blocks


def is_collapsed(block):
    try:
        collapse = block.find_element(
            by="class name", value="review-collapsed__content"
        )
        if collapse:
            return "被折叠"
    except:
        return ""
        pass


def get_user(block):
    try:
        user = block.find_element(
            by="class name", value="tap-router.user-name__text"
        ).get_attribute("title")

        return user
    except:
        return "无法获取用户名"
        pass


def get_score(block):
    try:
        style = block.find_element(
            by="class name", value="tap-rate__highlight"
        ).get_attribute("style")
        width = style[(style.index(": ") + 1):style.index("px")].strip()
        score = int(int(width) / 18)  # 每个星长度为18

        # print(score)
        return score
    except:
        return 0
        pass


def get_time_text(block):
    try:  # 获取确切时间
        time_text = block.find_element(
            by="class name", value="tap-time.review-item__updated-time"
        ).get_attribute("title")
    except:  # 获取最后修改时间
        try:
            time_text = block.find_element(
                by="class name", value="review-item__modified-tip").text
        # print(time)
        except:
            return "无法获取时间"
            pass
    return time_text


def get_time(post_time: str):
    now = datetime.datetime.now()
    null = "无法获取时间" in post_time
    if null:
        return "无法获取时间"
    delta = "于" in post_time
    if delta:
        delta_text = post_time[post_time.index("于") + 1:].strip()
        yesterday = "昨天" in post_time
        day_delta = "天" in post_time
        hour_delta = "小时" in post_time
        min_delta = "分钟" in post_time
        sec_delta = "秒" in post_time
        if yesterday:
            exact_time = now - datetime.timedelta(days=1)
        elif day_delta:
            delta_days = delta_text[:delta_text.index("天")].strip()
            exact_time = now - datetime.timedelta(days=int(delta_days))
        elif hour_delta:
            delta_hours = delta_text[:delta_text.index("小时")].strip()
            exact_time = now - datetime.timedelta(hours=int(delta_hours))
        elif min_delta:
            delta_min = delta_text[:delta_text.index("分钟")].strip()
            exact_time = now - datetime.timedelta(minutes=int(delta_min))
        elif sec_delta:
            delta_sec = delta_text[:delta_text.index("秒")].strip()
            exact_time = now - datetime.timedelta(minutes=int(delta_sec))
        else:
            exact_time = datetime.datetime.strptime(delta_text, "%Y/%m/%d")
    else:
        exact_time = datetime.datetime.strptime(post_time, "%Y/%m/%d %H:%M:%S")
    return exact_time


def get_comment(block):
    try:
        comment = block.find_element(
            by="class name", value="text-box__content"
        ).text

        return comment
    except:
        return "无法获取评论"
        pass


def Do():
    get_block(url_full)


def do_print_xl():
    print("正在写入文档")
    wb = xl.Workbook()
    result_sheet = wb.create_sheet("result", 0)
    count_sheet = wb.create_sheet("count", 1)

    result_sheet.append(label_R)
    for i in range(len(user_out)):
        data = [user_out[i], comment_out[i], score_out[i],
                time_out[i], collapse_out[i], date_out[i]]
        result_sheet.append(data)

    count_sheet.append(label_C)
    date_result = collections.Counter(date_out)
    # print(date_result)
    index_list = []
    index = 0
    for date, num in date_result.items():
        date_present.append(date)
        index_list.append(index)
        index += num  # 下标依照数量累加

    for i in range(len(date_present)):
        if 0 < i < len(date_present) - 1:  # 不满一天的忽略
            star_counter = score_out[index_list[i]:index_list[i + 1]]
            score_result = collections.Counter(star_counter)
            data = [date_present[i],
                    score_result[1], score_result[2], score_result[3], score_result[4], score_result[5]]
            count_sheet.append(data)

    wb.save(file_name)


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    Do()
    do_print_xl()
